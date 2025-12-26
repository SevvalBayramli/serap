from flask import Flask, render_template, request, send_file
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

LAST_FILE = None  # indirilecek son dosya

RENKLER = [
    "EEF2FF",  # açık mor
    "ECFEFF",  # açık mavi
    "FEF3C7",  # açık sarı
    "FCE7F3",  # açık pembe
    "DCFCE7",  # açık yeşil
]

@app.route("/", methods=["GET", "POST"])
def index():
    global LAST_FILE

    tables = None

    if request.method == "POST":
        file = request.files["file"]

        if "file" not in request.files:
            return "Dosya gelmedi"

        file = request.files["file"]

        if file.filename == "":
            return "Dosya seçilmedi"

        if not file or not file.filename.endswith(".xlsx"):
            return "Lütfen Excel (.xlsx) dosyası yükleyin"

        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)

        # === EXCEL İŞLEME ===
        dosya = filepath
        genel_sayfa_adi = "GENEL_TOPLAM"

        try:
            excel = pd.ExcelFile(dosya)
        except Exception as e:
            return f"Excel okunamadı: {str(e)}"
        tum_urunler = []

        for sheet in excel.sheet_names:
            if sheet == genel_sayfa_adi:
                continue

            df = excel.parse(sheet, header=None)

            urun_adi_row = urun_kod_row = toplam_row = None

            for i in range(len(df)):
                hucre = str(df.iloc[i, 0]).strip().upper()
                if hucre == "ÜRÜN ADI":
                    urun_adi_row = i
                elif hucre in ["ÜRÜN KOD", "ÜRÜN KODU"]:
                    urun_kod_row = i
                elif hucre == "TOPLAM":
                    toplam_row = i

            # === SATIR TESPİTİ ===
            if None in (urun_adi_row, urun_kod_row, toplam_row):
                fallback_mode = True
                urun_adi_row = 1   # 2. satır
                urun_kod_row = 2   # 3. satır
            else:
                fallback_mode = False

                
            for col in range(len(df.columns)):
                urun_adi = df.iloc[urun_adi_row, col]
                kodu = df.iloc[urun_kod_row, col]

                if pd.isna(urun_adi) or pd.isna(kodu):
                    continue

                # === NORMAL MOD (TOPLAM SATIRI VAR) ===
                if not fallback_mode:
                    toplam = df.iloc[toplam_row, col]
                    toplam_numeric = pd.to_numeric(toplam, errors="coerce")
                    if pd.isna(toplam_numeric):
                        continue

                # === FALLBACK MOD (ALTTAKİ SAYILARI TOPLA) ===
                else:
                    sayilar = df.iloc[urun_kod_row + 1 :, col]
                    sayilar_numeric = pd.to_numeric(sayilar, errors="coerce")
                    toplam_numeric = sayilar_numeric.sum()

                    if toplam_numeric == 0:
                        continue

                tum_urunler.append({
                    "SAYFA": sheet,
                    "ÜRÜN ADI": str(urun_adi).strip(),
                    "KODU": str(kodu).strip(),
                    "TOPLAM": float(toplam_numeric)
                })

        genel_df = pd.DataFrame(tum_urunler)

        with pd.ExcelWriter(
            dosya,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace"
        ) as writer:
            genel_df.to_excel(writer, sheet_name=genel_sayfa_adi, index=False)

        # === EXCEL RENKLENDİRME ===
        wb = load_workbook(dosya)
        ws = wb[genel_sayfa_adi]

        sayfa_renk_map = {}
        renk_index = 0

        # 1. satır başlık olduğu için 2'den başlıyoruz
        for row in range(2, ws.max_row + 1):
            sayfa_adi = ws.cell(row=row, column=1).value  # SAYFA kolonu

            if sayfa_adi not in sayfa_renk_map:
                sayfa_renk_map[sayfa_adi] = RENKLER[renk_index % len(RENKLER)]
                renk_index += 1

            renk = sayfa_renk_map[sayfa_adi]
            fill = PatternFill(start_color=renk, end_color=renk, fill_type="solid")

            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

        wb.save(dosya)


        LAST_FILE = dosya
        tables = genel_df.to_dict(orient="records")

    return render_template("index.html", tables=tables)


@app.route("/download")
def download():
    if LAST_FILE:
        return send_file(LAST_FILE, as_attachment=True)
    return "İndirilecek dosya yok"


if __name__ == "__main__":
    app.run(debug=True)
